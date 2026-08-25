
#!/usr/bin/env python3
"""Aplica revisões do dashboard e gera CSVs/XLSX de consulta.

Não altera a base reconciliada em paineis/: as saídas revisadas ficam em
data/ e preservam um histórico explícito.
"""
from __future__ import annotations
import csv, json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

ROOT=Path(__file__).resolve().parents[1]
DATA=ROOT/"data"
DASH=DATA/"dashboard.json"
REVISIONS=DATA/"revisoes_aprovadas.json"
OUT_ITEMS=DATA/"painel_itens.csv"
OUT_VOLUME=DATA/"painel_volumetria.csv"
OUT_XLSX=DATA/"acervo_hsia_atualizado.xlsx"
CATEGORIES=["colunas_autorais","entrevistas_escritas_completas","transcricoes_palestras_aulas",
 "participacoes_em_reportagens","casos_relatorios_profissionais","artigos_profissionais",
 "artigos_academicos","videos_podcasts","livros_capitulos"]
FIELDS=["fonte","categoria_painel","manifestacao_id","item_id","producao_principal_item_id",
 "manifestacao_derivada_de_item_id","papel_manifestacao","tipo_manifestacao","titulo","ano",
 "data_publicacao","tipo_publicacao","autores","papel_hsia","url_original","url_principal",
 "urls_secundarias","texto_ou_evidencia_disponivel","republicacao_de_item_id","status_verificacao",
 "contabilizado_na_volumetria","papeis_da_fonte","status_acesso","conteudo_status","evidencia",
 "rotulo_credito","credito_exibicao","participacao_hsia"]

def load(path,default):
 try:return json.loads(path.read_text(encoding="utf-8"))
 except (OSError,json.JSONDecodeError):return default
def active_rows(dashboard,revisions):
 removed={x.get("manifestacao_id",x) if isinstance(x,dict) else x for x in revisions.get("removed",[])}
 rows=[x for x in dashboard["itens"] if x.get("contabilizado",True) and x.get("manifestacao_id") not in removed]
 rows.extend(x for x in revisions.get("added",[]) if x.get("manifestacao_id") not in removed)
 clean=[]
 for row in rows:
  item={field:row.get(field,"") for field in FIELDS}
  if "desconhecid" in str(item.get("autores","")).lower():item["autores"]=""
  if not item.get("credito_exibicao"):
   category=item.get("categoria_painel","");source=item.get("fonte","Fonte não informada");authors=item.get("autores","")
   if category in {"colunas_autorais","artigos_profissionais","artigos_academicos","livros_capitulos"}:label,credit,role="Autoria",authors or "Hsia Hua Sheng","Autor ou coautor"
   elif category=="entrevistas_escritas_completas":label,credit,role="Reportagem ou entrevista",authors or f"Redação de {source}","Entrevistado: Hsia Hua Sheng"
   elif category=="participacoes_em_reportagens":label,credit,role="Reportagem",authors or f"Redação de {source}","Participação de Hsia Hua Sheng"
   elif category=="videos_podcasts":label,credit,role="Produção",authors or source,"Participação de Hsia Hua Sheng"
   else:label,credit,role="Crédito",authors or f"Responsabilidade editorial de {source}","Registro relacionado a Hsia Hua Sheng"
   item.update(rotulo_credito=label,credito_exibicao=credit,participacao_hsia=role)
  clean.append(item)
 return clean
def volumes(rows,dashboard):
 template={x.get("fonte"):x for x in dashboard.get("volumetria",[])}
 counts=Counter((x["fonte"],x["categoria_painel"]) for x in rows)
 sources=sorted({x["fonte"] for x in rows}|set(template),key=str.casefold);out=[]
 for source in sources:
  original=template.get(source,{})
  row={"fonte":source,"forma_de_acesso":original.get("forma_de_acesso","Link informado na revisão")}
  for category in CATEGORIES:row[category]=counts[source,category]
  row["total_producoes_unicas"]=sum(row[c] for c in CATEGORIES)
  row["pendentes"]=original.get("pendentes",0)
  row["status_reconciliacao"]=original.get("status_reconciliacao","incluido_por_revisao")
  out.append(row)
 return out
def write_csv(path,rows,fields):
 with path.open("w",encoding="utf-8-sig",newline="") as f:
  writer=csv.DictWriter(f,fieldnames=fields,extrasaction="ignore");writer.writeheader();writer.writerows(rows)
def write_xlsx(rows,volume,history):
 try:
  from openpyxl import Workbook
  from openpyxl.styles import Font, PatternFill
 except ImportError:raise SystemExit("Instale openpyxl para gerar o XLSX: py -m pip install openpyxl")
 wb=Workbook();ws=wb.active;ws.title="Publicações"
 def add(sheet,data,fields):
  sheet.append(fields)
  for row in data:sheet.append([row.get(k,"") for k in fields])
  sheet.freeze_panes="A2";sheet.auto_filter.ref=sheet.dimensions
  for cell in sheet[1]:
   cell.font=Font(name="Calibri",size=11,bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="0D2832")
  for col in sheet.columns:
   letter=col[0].column_letter;sheet.column_dimensions[letter].width=min(48,max(12,max(len(str(c.value or "")) for c in col)+2))
 add(ws,rows,FIELDS)
 vfields=["fonte","forma_de_acesso",*CATEGORIES,"total_producoes_unicas","pendentes","status_reconciliacao"]
 add(wb.create_sheet("Volumetria"),volume,vfields)
 hfields=["type","reviewer","changed_at","manifestacao_id","item_id","titulo","fonte"]
 flat=[]
 for entry in history:
  item=entry.get("item",{}) if isinstance(entry,dict) else {}
  flat.append({k:entry.get(k,item.get(k,"")) for k in hfields})
 add(wb.create_sheet("Histórico"),flat,hfields)
 wb.properties.title="Acervo Hsia Hua Sheng — revisão do dashboard"
 wb.properties.modified=datetime.now(timezone.utc).replace(tzinfo=None)
 wb.save(OUT_XLSX)

def main():
 dashboard=load(DASH,None)
 if not dashboard:raise SystemExit(f"Arquivo ausente ou inválido: {DASH}")
 revisions=load(REVISIONS,{"removed":[],"added":[],"history":[]})
 rows=active_rows(dashboard,revisions);volume=volumes(rows,dashboard)
 write_csv(OUT_ITEMS,rows,FIELDS)
 vfields=["fonte","forma_de_acesso",*CATEGORIES,"total_producoes_unicas","pendentes","status_reconciliacao"]
 write_csv(OUT_VOLUME,volume,vfields);write_xlsx(rows,volume,revisions.get("history",[]))
 print(f"{len(rows)} manifestações | {len(volume)} fontes | {OUT_XLSX.relative_to(ROOT)}")
if __name__=="__main__":main()
